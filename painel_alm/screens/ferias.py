import streamlit as st
import sys
import os
import io
from datetime import datetime
from collections import Counter

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'lib'))
from ferias_parser import (
    parse_ferias, extract_text_from_pdf, STATUS_LABEL, STATUS_COLOR,
    format_dias, build_whatsapp_texts,
)

st.title("📋 Programação de Férias")
st.caption(
    "Envie o PDF de 'Programação de Férias' exportado do Domínio Folha. "
    "O app classifica cada vínculo em Duplicidade / Vencida / Maturada dentro do prazo / A vencer, "
    "usando o campo 'Fer.' que o próprio Domínio calcula."
)

uploaded = st.file_uploader("Arraste ou selecione o PDF", type=["pdf"], key="ferias_upload")

if uploaded is not None:
    pdf_bytes = uploaded.read()

    with st.spinner("Lendo e classificando..."):
        text = extract_text_from_pdf(pdf_bytes)
        reference_date = datetime.now()
        results, audit, audit_ok, company_order = parse_ferias(text, reference_date=reference_date)

    if not results:
        st.error("Não consegui extrair nenhum registro deste PDF. Confere se é o relatório certo.")
        st.stop()

    if audit_ok:
        st.success(f"✅ Conferido: todas as {len(audit)} empresas batem com o 'Total de empregados' do próprio PDF.")
    else:
        st.warning("⚠️ Alguma empresa não bateu a contagem declarada no PDF.")
        with st.expander("Ver detalhes da auditoria"):
            for empresa, declarado, parseado in audit:
                if declarado != parseado:
                    st.write(f"❌ {empresa}: declarado={declarado}, parseado={parseado}")

    counts = Counter(r['status'] for r in results)
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("🔴 Duplicidade", counts.get('DUPLICIDADE', 0))
    c2.metric("🟠 Vencida", counts.get('VENCIDA', 0))
    c3.metric("🟡 Maturada (dentro do prazo)", counts.get('MATURADA', 0))
    c4.metric("⚪ A vencer", counts.get('A_VENCER', 0))

    st.divider()

    whatsapp_texts = build_whatsapp_texts(results, company_order)
    if whatsapp_texts:
        st.subheader("📲 Texto para WhatsApp (empregados com 30 dias)")
        st.caption("Um bloco por empresa, na ordem em que aparecem no PDF. Só entram aqui quem tem direito a período cheio (30 dias).")
        for empresa in company_order:
            if empresa not in whatsapp_texts:
                continue
            st.markdown(f"**{empresa}**")
            st.code(whatsapp_texts[empresa], language=None)
        st.divider()

    status_options = ["Todos"] + [STATUS_LABEL[s] for s in ["DUPLICIDADE", "VENCIDA", "MATURADA", "A_VENCER"]]
    filtro = st.selectbox("Filtrar por status", status_options)
    filtered = results if filtro == "Todos" else [r for r in results if r['status_label'] == filtro]

    st.write(f"**{len(filtered)} registros** (de {len(results)} no total)")

    # Blocos por empresa, na ordem em que aparecem no PDF (não em ordem alfabética
    # nem por urgência) - dentro de cada empresa, mantém a ordem de urgência que
    # já vem em 'filtered'. Sem cor de fundo no nome da empresa, só nas linhas.
    by_empresa = {}
    for r in filtered:
        by_empresa.setdefault(r['empresa'], []).append(r)

    for empresa in company_order:
        rows = by_empresa.get(empresa)
        if not rows:
            continue
        st.markdown(f"#### {empresa}")
        table_html = "<table style='width:100%; border-collapse: collapse;'>"
        table_html += (
            "<tr style='font-weight:bold; border-bottom: 2px solid #333;'>"
            "<td style='padding:6px;'>Código</td><td style='padding:6px;'>Nome</td>"
            "<td style='padding:6px;'>Status</td><td style='padding:6px;'>Dias</td>"
            "<td style='padding:6px;'>Limite p/ gozo</td></tr>"
        )
        for r in rows:
            lim = r['limite'].strftime('%d/%m/%Y') if r['limite'] else '-'
            color = STATUS_COLOR[r['status']]
            table_html += (
                f"<tr style='background-color:#{color}; border-bottom:1px solid #ddd;'>"
                f"<td style='padding:6px;'>{r['codigo']}</td><td style='padding:6px;'>{r['nome']}</td>"
                f"<td style='padding:6px;'>{r['status_label']}</td><td style='padding:6px;'>{format_dias(r['dias'])}</td>"
                f"<td style='padding:6px;'>{lim}</td></tr>"
            )
        table_html += "</table>"
        st.markdown(table_html, unsafe_allow_html=True)
        st.write("")

    st.divider()

    def build_excel(rows):
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font
        wb = Workbook()
        ws = wb.active
        ws.title = "Ferias"
        headers = ["Empresa", "Código", "Nome", "Status", "Dias", "Limite p/ gozo"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            lim = r['limite'].strftime('%d/%m/%Y') if r['limite'] else '-'
            ws.append([r['empresa'], r['codigo'], r['nome'], r['status_label'], format_dias(r['dias']), lim])
            fill = PatternFill(start_color=STATUS_COLOR[r['status']], end_color=STATUS_COLOR[r['status']], fill_type="solid")
            for cell in ws[ws.max_row]:
                cell.fill = fill
        for col_cells in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(length + 2, 50)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    excel_buf = build_excel(filtered)
    st.download_button(
        "📥 Baixar Excel colorido",
        data=excel_buf,
        file_name=f"ferias_classificado_{datetime.now().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
