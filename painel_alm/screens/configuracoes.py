import streamlit as st
import sys
import os
import re
import pandas as pd

sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'lib'))
import db

st.title("⚙️ Configurações — Lista de clientes")
st.caption(
    "Essa é a lista que aparece no Dashboard e na busca. Edita direto na tabela "
    "(clica numa célula pra mudar, usa o + no final pra adicionar linha, ou o "
    "ícone de lixeira pra apagar), depois clica em Salvar."
)

clientes = db.listar_clientes()
df = pd.DataFrame(clientes) if clientes else pd.DataFrame(columns=['codigo', 'empresa', 'nome', 'cnpj'])

st.subheader("📋 Tabela de clientes")
edited_df = st.data_editor(
    df,
    num_rows="dynamic",
    use_container_width=True,
    column_config={
        'codigo': st.column_config.NumberColumn("Código", width="small"),
        'empresa': st.column_config.TextColumn("Empresa", width="medium", required=True),
        'nome': st.column_config.TextColumn("Nome/contato", width="medium"),
        'cnpj': st.column_config.TextColumn("CNPJ/CPF", width="medium", required=True),
    },
    key="clientes_editor",
)

if st.button("💾 Salvar alterações"):
    lista = edited_df.to_dict('records')
    # remove linhas vazias (ex: linha nova sem preencher ainda)
    lista = [c for c in lista if str(c.get('empresa') or '').strip() and str(c.get('cnpj') or '').strip()]
    db.salvar_clientes(lista)
    st.success(f"Salvo! {len(lista)} clientes na lista.")
    st.rerun()

st.divider()
st.subheader("📥 Importar de uma planilha (substitui a lista toda)")
st.caption(
    "Use isso pra trocar a lista inteira de uma vez (ex: começar um painel novo pra outro escritório). "
    "A planilha precisa ter colunas: codigo, empresa, nome, cnpj."
)
upload = st.file_uploader("Planilha (.xlsx)", type=["xlsx"], key="import_clientes")
if upload is not None:
    import openpyxl
    wb = openpyxl.load_workbook(upload, data_only=True)
    ws = wb.worksheets[0]
    header = [str(c.value or '').strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    lista_importada = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(header, row))
        empresa = str(row_dict.get('empresa') or '').strip()
        cnpj = str(row_dict.get('cnpj') or '').strip()
        if not empresa or not cnpj:
            continue
        lista_importada.append({
            'codigo': row_dict.get('codigo'),
            'empresa': empresa,
            'nome': str(row_dict.get('nome') or '').strip(),
            'cnpj': cnpj,
        })
    st.write(f"{len(lista_importada)} clientes encontrados na planilha.")
    if st.button("⚠️ Confirmar e substituir lista atual"):
        db.salvar_clientes(lista_importada)
        st.success("Lista substituída!")
        st.rerun()
